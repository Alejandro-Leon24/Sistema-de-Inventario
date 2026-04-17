import pytest
from datetime import datetime
from openpyxl import Workbook

from app.app import app as flask_app, BASE_DIR
from app.blueprints.inventory import (
    _build_mapped_chunk,
    _build_row_key,
    _load_excel_with_detected_headers,
    _map_paste_row_best_effort,
    _normalize_row_for_exact_compare,
)
from database.controller import _normalize_inventory_date_for_output
from database.schema_manager import init_schema


@pytest.fixture()
def client(tmp_path):
    db_path = tmp_path / "test_inventario.db"
    flask_app.config["TESTING"] = True
    flask_app.config["DATABASE"] = db_path

    with flask_app.app_context():
        init_schema(BASE_DIR)

    with flask_app.test_client() as test_client:
        yield test_client


def test_create_and_list_estados(client):
    payload = {"nombre": "Estado QA", "descripcion": "Creado por pytest"}
    create_response = client.post("/api/parametros/estados", json=payload)
    assert create_response.status_code in (201, 409)

    list_response = client.get("/api/parametros/estados")
    assert list_response.status_code == 200
    data = list_response.get_json()
    assert "data" in data
    assert isinstance(data["data"], list)
    assert any((row.get("nombre") or "").lower() == "estado qa" for row in data["data"])


def test_create_and_list_condiciones(client):
    payload = {"nombre": "Condicion QA", "descripcion": "Creado por pytest"}
    create_response = client.post("/api/parametros/condiciones", json=payload)
    assert create_response.status_code in (201, 409)

    list_response = client.get("/api/parametros/condiciones")
    assert list_response.status_code == 200
    data = list_response.get_json()
    assert "data" in data
    assert isinstance(data["data"], list)
    assert any((row.get("nombre") or "").lower() == "condicion qa" for row in data["data"])


def test_excel_loader_ignora_filas_vacias(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Cod Inv", "Descripcion", "Marca"])
    sheet.append(["", "", ""])
    sheet.append(["   ", " - ", "   "])
    sheet.append(["INV-001", "Laptop", "Dell"])

    file_path = tmp_path / "import_vacias.xlsx"
    workbook.save(file_path)

    parsed = _load_excel_with_detected_headers(file_path)
    assert parsed["headers"]
    assert len(parsed["data_rows"]) == 1
    assert parsed["data_rows"][0][0] == "INV-001"


def test_excel_loader_rellena_celdas_combinadas(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Cod Inv", "Descripcion", "Marca"])
    sheet.append(["INV-100", "Monitor 1", "HP"])
    sheet.append(["", "Monitor 2", "HP"])
    sheet.merge_cells("A2:A3")

    file_path = tmp_path / "import_merged.xlsx"
    workbook.save(file_path)

    parsed = _load_excel_with_detected_headers(file_path)
    assert len(parsed["data_rows"]) == 2
    assert parsed["data_rows"][0][0] == "INV-100"
    assert parsed["data_rows"][1][0] == "INV-100"


def test_build_mapped_chunk_descarta_fila_totalmente_vacia():
    data_rows = [
        ["", "", ""],
        ["INV-200", "Impresora", ""],
    ]
    mapping = {
        "0": "cod_inventario",
        "1": "descripcion",
        "2": "marca",
    }

    chunk = _build_mapped_chunk(data_rows, mapping, start_index=0, chunk_size=20)

    assert len(chunk) == 1
    assert chunk[0]["row_index"] == 1
    assert chunk[0]["data"]["cod_inventario"] == "INV-200"


def test_build_mapped_chunk_descarta_placeholders_sin_datos_reales():
    data_rows = [
        ["S/C", " - ", "N/A", ""],
        ["S/C", "Computador", "Dell", "Lab 1"],
    ]
    mapping = {
        "0": "cod_inventario",
        "1": "descripcion",
        "2": "marca",
        "3": "ubicacion",
    }

    chunk = _build_mapped_chunk(data_rows, mapping, start_index=0, chunk_size=20)

    assert len(chunk) == 1
    assert chunk[0]["row_index"] == 1
    assert chunk[0]["data"]["descripcion"] == "Computador"


def test_excel_loader_fecha_datetime_solo_fecha(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Fecha Adquisicion", "Descripcion"])
    sheet.append([datetime(2026, 4, 17, 15, 45, 30), "Equipo QA"])

    file_path = tmp_path / "import_fechas_datetime.xlsx"
    workbook.save(file_path)

    parsed = _load_excel_with_detected_headers(file_path)
    assert len(parsed["data_rows"]) == 1
    assert parsed["data_rows"][0][0] == "2026-04-17"


def test_normalizacion_compare_hace_match_exacto_con_mismo_excel():
    imported_row = {
        "cod_inventario": "INV-300",
        "cod_esbye": "ESB-900",
        "cantidad": "1.0",
        "descripcion": "Proyector QA",
        "ubicacion": "Bloque A / Piso 1 / Lab 1",
        "marca": "Epson",
        "modelo": "X123",
        "serie": "SER-777",
        "estado": "Bueno",
        "usuario_final": "Juan Perez",
        "fecha_adquisicion": "17/04/2026 00:00:00",
        "valor": "1.200,00",
        "observacion": "Sin novedades",
    }

    db_like_row = {
        "cod_inventario": "INV-300",
        "cod_esbye": "ESB-900",
        "cantidad": 1,
        "descripcion": "Proyector QA",
        "ubicacion": "Bloque A / Piso 1 / Lab 1",
        "marca": "Epson",
        "modelo": "X123",
        "serie": "SER-777",
        "estado": "Bueno",
        "usuario_final": "Juan Perez",
        "fecha_adquisicion": "2026-04-17",
        "valor": 1200.0,
        "observacion": "Sin novedades",
    }

    imported_key = _build_row_key(_normalize_row_for_exact_compare(imported_row))
    db_key = _build_row_key(_normalize_row_for_exact_compare(db_like_row))

    assert imported_key == db_key


def test_map_paste_row_best_effort_detecta_con_offset_alto():
    raw_row = ["", "", "", "", "INV-550", "ESB-550", "Cuenta QA", "1", "Router", "Cisco", "R1", "SER-1", "Bueno", "Bloque A / Piso 1 / Lab", "2026-04-17", "1000", "Juan Perez"]

    mapped = _map_paste_row_best_effort(raw_row)

    assert mapped.get("cod_inventario") == "INV-550"
    assert mapped.get("descripcion") == "Router"
    assert mapped.get("ubicacion") == "Bloque A / Piso 1 / Lab"


def test_normalizacion_compare_usuario_final_con_titulo_match_exacto():
    imported_row = {
        "cod_inventario": "INV-901",
        "descripcion": "Switch",
        "usuario_final": "Ing. Juan Perez",
        "cantidad": "1",
        "fecha_adquisicion": "2026-04-17",
    }

    db_like_row = {
        "cod_inventario": "INV-901",
        "descripcion": "Switch",
        "usuario_final": "Juan Perez",
        "cantidad": 1,
        "fecha_adquisicion": "2026-04-17",
    }

    imported_key = _build_row_key(_normalize_row_for_exact_compare(imported_row))
    db_key = _build_row_key(_normalize_row_for_exact_compare(db_like_row))

    assert imported_key == db_key


def test_normalizacion_compare_cantidad_vacia_equivale_a_uno():
    imported_row = {
        "cod_inventario": "INV-902",
        "descripcion": "Access Point",
        "cantidad": "",
        "fecha_adquisicion": "2026-04-17",
    }

    db_like_row = {
        "cod_inventario": "INV-902",
        "descripcion": "Access Point",
        "cantidad": 1,
        "fecha_adquisicion": "2026-04-17",
    }

    imported_key = _build_row_key(_normalize_row_for_exact_compare(imported_row))
    db_key = _build_row_key(_normalize_row_for_exact_compare(db_like_row))

    assert imported_key == db_key


def test_normalize_inventory_date_for_output_iso_datetime():
    assert _normalize_inventory_date_for_output("2026-04-17 00:00:00") == "2026-04-17"


def test_normalize_inventory_date_for_output_dmy_datetime():
    assert _normalize_inventory_date_for_output("17/04/2026 00:00:00") == "2026-04-17"
