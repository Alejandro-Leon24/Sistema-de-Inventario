from io import BytesIO
from datetime import datetime

def generar_excel(items, columnas, titulo_hoja):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError('Para exportar a Excel se requiere openpyxl instalado.')

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = titulo_hoja

    headers = [label for _, label in columnas]
    worksheet.append(headers)

    header_fill = PatternFill(fill_type="solid", start_color="D9EEF9", end_color="D9EEF9")
    header_font = Font(bold=True, color="1F2937")
    
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_item in items:
        row_values = []
        for field, _ in columnas:
            try:
                value = row_item[field]
            except (KeyError, IndexError):
                value = ""
                
            if hasattr(row_item, 'get'):
                value = row_item.get(field, value)
                
            row_values.append("" if value is None else value)
        worksheet.append(row_values)

    worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(headers)).column_letter}{worksheet.max_row}"
    worksheet.freeze_panes = "A2"

    for column_cells in worksheet.columns:
        max_len = 0
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(cell_value))
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_len + 2, 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output