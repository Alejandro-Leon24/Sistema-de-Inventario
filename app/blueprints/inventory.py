import sqlite3
import logging
import json
import os
import re
import tempfile
import time
import uuid
import unicodedata
from datetime import date, datetime, timedelta

from flask import Blueprint, jsonify, request
from database.db import get_db
from database.controller import resolve_or_create_personal_name

from database.inventory_repository import (
    ALLOWED_INVENTORY_FIELDS,
    bulk_insert_inventory_dicts,
    clear_inventory_items,
    create_inventory_item,
    delete_inventory_item,
    find_inventory_code_duplicates,
    get_column_mappings,
    get_inventory_item,
    get_inventory_search_diagnostics,
    list_inventory_items_paginated,
    replace_column_mappings,
    set_user_preference,
    get_user_preferences,
    update_inventory_item,
)


inventory_bp = Blueprint("inventory", __name__)
DEFAULT_USER_KEY = "portable_user"
logger = logging.getLogger(__name__)

_EXCEL_IMPORT_DIR = os.path.join(tempfile.gettempdir(), "inventario_excel_import")
os.makedirs(_EXCEL_IMPORT_DIR, exist_ok=True)

_MAX_EXCEL_PREVIEW_ROWS = 20
_MAX_EXCEL_IMPORT_ROWS = 10_000
_MAX_EXCEL_FILE_BYTES = 10 * 1024 * 1024
_DEFAULT_IMPORT_CHUNK_SIZE = 20
_UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
    re.IGNORECASE,
)

_IMPORT_CANONICAL_FIELDS = [
    "cod_inventario",
    "cod_esbye",
    "cuenta",
    "cantidad",
    "descripcion",
    "ubicacion",
    "marca",
    "modelo",
    "serie",
    "estado",
    "condicion",
    "usuario_final",
    "fecha_adquisicion",
    "valor",
    "observacion",
    "descripcion_esbye",
    "marca_esbye",
    "modelo_esbye",
    "serie_esbye",
    "fecha_adquisicion_esbye",
    "valor_esbye",
    "ubicacion_esbye",
    "observacion_esbye",
]

# Orden real del pegado rapido desde Excel en Inventario (tabla/formulario).
_PASTE_CANONICAL_FIELDS = [
    "cod_inventario",
    "cod_esbye",
    "cuenta",
    "cantidad",
    "descripcion",
    "marca",
    "modelo",
    "serie",
    "estado",
    "ubicacion",
    "fecha_adquisicion",
    "valor",
    "usuario_final",
    "observacion",
    "descripcion_esbye",
    "marca_esbye",
    "modelo_esbye",
    "serie_esbye",
    "fecha_adquisicion_esbye",
    "valor_esbye",
    "ubicacion_esbye",
    "observacion_esbye",
]

_EMPTY_PLACEHOLDER_NORMALIZED = {
    "na",
    "n a",
    "nd",
    "n d",
    "sd",
    "s d",
    "null",
    "none",
    "ninguno",
    "sin dato",
    "sin datos",
    "sin informacion",
    "sindato",
    "sindatos",
    "sininformacion",
    "sininfo",
    "vacio",
}

_SIGNIFICANT_IMPORT_FIELDS = {
    "cod_inventario",
    "cod_esbye",
    "cuenta",
    "descripcion",
    "marca",
    "modelo",
    "serie",
    "estado",
    "ubicacion",
    "usuario_final",
    "observacion",
    "descripcion_esbye",
    "marca_esbye",
    "modelo_esbye",
    "serie_esbye",
    "ubicacion_esbye",
    "observacion_esbye",
}


def _looks_like_person_name(value):
    text = str(value or "").strip()
    if not text:
        return False
    tokens = [tok for tok in re.split(r"\s+", text) if tok]
    alpha_tokens = [tok for tok in tokens if re.search(r"[A-Za-zÁÉÍÓÚáéíóúÑñ]", tok)]
    if len(alpha_tokens) < 2:
        return False
    lowered = _normalize_text(text)
    non_person_markers = {"s m", "s s", "n a", "na", "ninguno", "sin datos"}
    return lowered not in non_person_markers


def _looks_like_date_text(value):
    text = str(value or "").strip()
    if not text:
        return False
    return bool(
        re.fullmatch(r"\d{4}-\d{2}-\d{2}", text)
        or re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text)
        or re.fullmatch(r"\d+(?:\.\d+)?", text)
    )


def _looks_like_money_text(value):
    text = str(value or "").strip().replace(" ", "")
    if not text:
        return False
    return bool(re.fullmatch(r"[-+]?\d{1,3}(?:[\.,]\d{3})*(?:[\.,]\d+)?|[-+]?\d+(?:[\.,]\d+)?", text))


def _score_paste_mapping(mapped):
    score = 0
    if _looks_like_person_name(mapped.get("usuario_final")):
        score += 4
    elif mapped.get("usuario_final") not in (None, ""):
        score -= 3

    cantidad_text = str(mapped.get("cantidad") or "").strip()
    if re.fullmatch(r"\d+", cantidad_text):
        cantidad_num = int(cantidad_text)
        if 0 < cantidad_num <= 1000:
            score += 2
    elif cantidad_text:
        score -= 1

    fecha_text = str(mapped.get("fecha_adquisicion") or "").strip()
    if _looks_like_date_text(fecha_text):
        score += 2
    elif fecha_text:
        score -= 2

    valor_text = str(mapped.get("valor") or "").strip()
    if _looks_like_money_text(valor_text):
        score += 2
    elif valor_text:
        score -= 2
    ubicacion = str(mapped.get("ubicacion") or "").strip().lower()
    if ubicacion and ("/" in ubicacion or "bloque" in ubicacion or "piso" in ubicacion or "area" in ubicacion):
        score += 2
    if str(mapped.get("cod_inventario") or "").strip():
        score += 1
    if str(mapped.get("descripcion") or "").strip():
        score += 1
    return score


def _map_paste_row_best_effort(raw_row):
    cells = list(raw_row or [])
    if not cells:
        return {}

    candidate_orders = [_PASTE_CANONICAL_FIELDS, _IMPORT_CANONICAL_FIELDS]
    best_row = {}
    best_score = float("-inf")
    best_offset = 0

    for order in candidate_orders:
        # Permite detectar filas con desplazamientos mayores (copias con columnas previas vacias).
        max_offset = min(8, max(len(cells) - 1, 0))
        possible_offsets = list(range(0, max_offset + 1))

        for offset in possible_offsets:
            mapped = {}
            for idx, field in enumerate(order):
                src_idx = idx + offset
                if src_idx >= len(cells):
                    break
                value = cells[src_idx]
                if isinstance(value, str):
                    value = value.strip()
                if value in (None, ""):
                    continue
                mapped[field] = value

            non_empty_count = len(mapped)
            key_field_hits = sum(1 for field in ("cod_inventario", "descripcion", "ubicacion") if mapped.get(field))
            score = _score_paste_mapping(mapped) + (non_empty_count * 0.35) + (key_field_hits * 1.25)

            # Preferimos menor offset cuando el score es similar para conservar el mapeo base esperado.
            if score > best_score or (score == best_score and offset < best_offset):
                best_score = score
                best_row = mapped
                best_offset = offset

    return best_row

_EXACT_ROW_COMPARE_FIELDS = [
    "cod_inventario",
    "cod_esbye",
    "cuenta",
    "cantidad",
    "descripcion",
    "ubicacion",
    "marca",
    "modelo",
    "serie",
    "estado",
    "usuario_final",
    "fecha_adquisicion",
    "valor",
    "observacion",
]

_HEADER_ALIAS_TO_FIELD = {
    "codigo inv": "cod_inventario",
    "codigo inventario": "cod_inventario",
    "cod inventario": "cod_inventario",
    "cod inv": "cod_inventario",
    "cod esbye": "cod_esbye",
    "codigo esbye": "cod_esbye",
    "esbye": "cod_esbye",
    "cuenta": "cuenta",
    "cant": "cantidad",
    "cantidad": "cantidad",
    "descripcion": "descripcion",
    "detalle": "descripcion",
    "marca": "marca",
    "modelo": "modelo",
    "serie": "serie",
    "estado": "estado",
    "condicion": "condicion",
    "ubicacion": "ubicacion",
    "fecha de adquisicion": "fecha_adquisicion",
    "fecha adquisicion": "fecha_adquisicion",
    "fecha compra": "fecha_adquisicion",
    "fecha": "fecha_adquisicion",
    "valor": "valor",
    "usuario final": "usuario_final",
    "custodio": "usuario_final",
    "observacion": "observacion",
    "observaciones": "observacion",
    "descripcion esbye": "descripcion_esbye",
    "marca esbye": "marca_esbye",
    "modelo esbye": "modelo_esbye",
    "serie esbye": "serie_esbye",
    "fecha esbye": "fecha_adquisicion_esbye",
    "fecha adq esbye": "fecha_adquisicion_esbye",
    "valor esbye": "valor_esbye",
    "ubicacion esbye": "ubicacion_esbye",
    "observacion esbye": "observacion_esbye",
}

_BASE_TO_ESBYE_FIELD = {
    "descripcion": "descripcion_esbye",
    "marca": "marca_esbye",
    "modelo": "modelo_esbye",
    "serie": "serie_esbye",
    "fecha_adquisicion": "fecha_adquisicion_esbye",
    "valor": "valor_esbye",
    "ubicacion": "ubicacion_esbye",
    "observacion": "observacion_esbye",
}


def _normalize_code_value(value):
    text = str(value or "").strip()
    if not text:
        return "S/C"

    compact = re.sub(r"[^a-z0-9]", "", text.lower())
    if compact in {"sc", "sincodigo", "sincod"}:
        return "S/C"
    return text


def _normalize_code_compare_value(value):
    normalized = _normalize_compare_value(_normalize_code_value(value))
    return "" if normalized == "s c" else normalized


def _normalize_codes_in_row_data(row_data):
    if not isinstance(row_data, dict):
        return row_data
    row_data["cod_inventario"] = _normalize_code_value(row_data.get("cod_inventario"))
    row_data["cod_esbye"] = _normalize_code_value(row_data.get("cod_esbye"))
    return row_data


def _normalize_text(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return " ".join(text.split())


def _looks_like_inventory_code_header(normalized_header):
    normalized = str(normalized_header or "").strip()
    if not normalized:
        return False
    if "esbye" in normalized:
        return False

    compact = normalized.replace(" ", "")
    if re.search(r"\bcod(?:igo)?\s*(inv|inventario)\b", normalized):
        return True
    if re.search(r"\bcod(?:igo)?\s*(jaf|prov)\b", normalized):
        return True
    if "codjaf" in compact or "codprov" in compact:
        return True

    tokens = set(normalized.split())
    has_code_word = "cod" in tokens or "codigo" in tokens
    has_jaf_or_prov = "jaf" in tokens or "prov" in tokens
    return has_code_word and has_jaf_or_prov


def _header_to_canonical_field(header_value):
    normalized = _normalize_text(header_value)
    if not normalized:
        return ""
    if _looks_like_inventory_code_header(normalized):
        return "cod_inventario"
    if normalized in _HEADER_ALIAS_TO_FIELD:
        return _HEADER_ALIAS_TO_FIELD[normalized]
    for alias, canonical in _HEADER_ALIAS_TO_FIELD.items():
        if normalized in alias or alias in normalized:
            return canonical
    return ""


def _normalize_suggested_mapping_by_repetition(headers, suggested_mapping):
    normalized_headers = [_normalize_text(header) for header in (headers or [])]
    result = []
    seen_base_count = {}

    for idx, raw_field in enumerate(suggested_mapping or []):
        field = str(raw_field or "").strip()
        header_norm = normalized_headers[idx] if idx < len(normalized_headers) else ""

        if field in _BASE_TO_ESBYE_FIELD:
            base_field = field
            seen_count = seen_base_count.get(base_field, 0)
            # Si el encabezado menciona ESBYE o es la segunda repetición del mismo campo,
            # mapear automáticamente a su variante ESBYE.
            if "esbye" in header_norm or seen_count >= 1:
                field = _BASE_TO_ESBYE_FIELD[base_field]
            seen_base_count[base_field] = seen_count + 1
        result.append(field)

    return result


def _header_row_score(row_values):
    used_fields = set()
    score = 0
    for value in row_values:
        canonical = _header_to_canonical_field(value)
        if canonical and canonical not in used_fields:
            used_fields.add(canonical)
            score += 1
    return score


def _looks_like_secondary_header_row(row_values):
    score = _header_row_score(row_values)
    if score < 2:
        return False

    non_empty = [str(value or "").strip() for value in row_values if str(value or "").strip()]
    if not non_empty:
        return False

    numeric_cells = 0
    for value in non_empty:
        compact = re.sub(r"\s+", "", value)
        if re.fullmatch(r"[\d.,/-]+", compact):
            numeric_cells += 1

    # Un encabezado secundario suele ser mayormente textual.
    return numeric_cells == 0


def _build_merged_cell_value_map(worksheet):
    merged_map = {}
    merged_ranges = getattr(getattr(worksheet, "merged_cells", None), "ranges", [])
    for merged_range in merged_ranges:
        try:
            anchor_value = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        except Exception:
            anchor_value = None

        # Si la celda ancla no tiene valor, no propagamos nada.
        if anchor_value in (None, ""):
            continue

        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(row_idx, col_idx)] = anchor_value

    return merged_map


def _has_meaningful_cell_value(value):
    text = str(value or "").strip()
    if not text:
        return False
    return bool(re.search(r"[0-9A-Za-zÁÉÍÓÚáéíóúÑñ]", text))


def _is_empty_placeholder_text(value):
    text = str(value or "").strip()
    if not text:
        return True

    normalized = _normalize_text(text)
    if not normalized:
        return True
    if normalized in _EMPTY_PLACEHOLDER_NORMALIZED:
        return True

    compact = re.sub(r"[^a-z0-9]", "", normalized)
    if compact in {"sc", "sincodigo", "sincod", "na", "nd", "sd", "vacio"}:
        return True

    return False


def _row_has_meaningful_import_data(row_dict):
    if not isinstance(row_dict, dict) or not row_dict:
        return False

    for field in _SIGNIFICANT_IMPORT_FIELDS:
        if field not in row_dict:
            continue
        raw_value = row_dict.get(field)
        if raw_value in (None, ""):
            continue
        if _is_empty_placeholder_text(raw_value):
            continue

        text = str(raw_value).strip()
        if re.search(r"[0-9A-Za-zÁÉÍÓÚáéíóúÑñ]", text):
            return True

    return False


def _load_excel_with_detected_headers(file_path):
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return {
            "header_row_index": 0,
            "headers": [],
            "data_rows": [],
        }
    merged_map = _build_merged_cell_value_map(ws)
    raw_rows = []
    max_cols = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        text_row = []
        for col_idx, cell in enumerate(row, start=1):
            value = cell
            if value in (None, ""):
                value = merged_map.get((row_idx, col_idx))

            if isinstance(value, datetime):
                value = value.date().isoformat()
            elif isinstance(value, date):
                value = value.isoformat()

            text_row.append(str(value).strip() if value is not None else "")
        raw_rows.append(text_row)
        if len(text_row) > max_cols:
            max_cols = len(text_row)
    wb.close()

    if not raw_rows or max_cols == 0:
        return {
            "header_row_index": 0,
            "headers": [],
            "data_rows": [],
        }

    rows = []
    for row in raw_rows:
        if len(row) < max_cols:
            row = row + [""] * (max_cols - len(row))
        rows.append(row)

    inspect_limit = min(len(rows), 60)
    best_idx = 0
    best_score = -1
    for idx in range(inspect_limit):
        current_score = _header_row_score(rows[idx])
        if current_score > best_score:
            best_score = current_score
            best_idx = idx

    if best_score < 1:
        best_idx = 0

    header_row = rows[best_idx]
    headers = [value if value else f"(col {idx + 1})" for idx, value in enumerate(header_row)]

    data_rows = []
    for idx in range(best_idx + 1, len(rows)):
        row = rows[idx]
        if not any(_has_meaningful_cell_value(cell) for cell in row):
            continue
        # Evita tomar como datos filas que parecen nuevos encabezados de otra tabla.
        if _looks_like_secondary_header_row(row):
            continue
        data_rows.append(row)
        if len(data_rows) >= _MAX_EXCEL_IMPORT_ROWS:
            break

    return {
        "header_row_index": best_idx,
        "headers": headers,
        "data_rows": data_rows,
    }


def _coerce_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _build_bulk_excel_procedencia_text(base_date=None):
    dt = base_date or datetime.now()
    date_text = dt.strftime("%d/%m/%Y %H:%M:%S")
    return f"Exportación Masiva de Excel - {date_text} / Bienes propios de la facultad"


def _normalize_compare_value(value):
    text = str(value if value is not None else "").strip()
    if text == "":
        return ""
    return _normalize_text(text)


def _normalize_person_name_for_compare(value):
    normalized = _normalize_text(value)
    if not normalized:
        return ""

    tokens = normalized.split()
    prefixes = {
        "ing", "ingeniero", "ingeniera", "dr", "dra", "doctor", "doctora",
        "lic", "licenciado", "licenciada", "abg", "abogada", "abogado",
        "arq", "arquitecto", "arquitecta", "tec", "tecnico", "tecnica",
        "sr", "sra", "srta", "msc", "mg", "mgs", "mgtr", "mtr", "mts", "mtro", "mt",
        "mrt", "prof", "profa", "tlgo", "tlga", "ts", "phd",
    }
    while tokens and tokens[0] in prefixes:
        tokens.pop(0)
    return " ".join(tokens)


def _normalize_import_date_for_compare(raw_value):
    if raw_value in (None, ""):
        return None

    if isinstance(raw_value, datetime):
        return raw_value.date().isoformat()
    if isinstance(raw_value, date):
        return raw_value.isoformat()

    text = str(raw_value).strip()
    if not text:
        return None

    iso_match = re.match(r"^(\d{4}-\d{2}-\d{2})", text)
    if iso_match:
        return iso_match.group(1)

    # Serial de Excel en texto numerico.
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        serial = float(text)
        if serial > 59:
            epoch = datetime(1899, 12, 30)
            parsed = epoch + timedelta(days=int(serial))
            return parsed.date().isoformat()

    patterns = [
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M:%S",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
    ]
    for pattern in patterns:
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue

    return text


def _normalize_int_like_value_for_compare(value):
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        number = float(text.replace(",", "."))
    except (TypeError, ValueError):
        return value
    if number.is_integer():
        return int(number)
    return value


def _normalize_money_value_for_compare(value):
    if value in (None, ""):
        return None
    text = str(value).strip().replace(" ", "")
    if not text:
        return None

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except (TypeError, ValueError):
        return value


def _normalize_row_for_exact_compare(row_data):
    normalized = dict(row_data or {})
    _normalize_codes_in_row_data(normalized)

    normalized["usuario_final"] = _normalize_person_name_for_compare(normalized.get("usuario_final"))

    normalized["cantidad"] = _normalize_int_like_value_for_compare(normalized.get("cantidad"))
    if normalized.get("cantidad") in (None, ""):
        # En inserción la cantidad vacía se guarda como 1; igualamos criterio para exact match.
        normalized["cantidad"] = 1

    for money_field in ("valor", "valor_esbye"):
        normalized[money_field] = _normalize_money_value_for_compare(normalized.get(money_field))

    for date_field in ("fecha_adquisicion", "fecha_adquisicion_esbye"):
        normalized[date_field] = _normalize_import_date_for_compare(normalized.get(date_field))

    return normalized


def _get_row_value(container, field):
    if isinstance(container, dict):
        return container.get(field)
    try:
        return container[field]
    except Exception:
        return None


def _build_row_key(row_dict):
    return tuple(_normalize_compare_value(_get_row_value(row_dict, field)) for field in _EXACT_ROW_COMPARE_FIELDS)


def _is_empty_row_key(key):
    return not any(key)


def _inventory_summary_item(row):
    return {
        "id": row["id"],
        "item_numero": row["item_numero"],
        "cod_inventario": row["cod_inventario"],
        "cod_esbye": row["cod_esbye"],
        "descripcion": row["descripcion"],
        "marca": row["marca"],
        "modelo": row["modelo"],
        "serie": row["serie"],
        "ubicacion": row["ubicacion"],
        "usuario_final": row["usuario_final"],
    }


def _build_match_fields(imported_row, db_summary):
    matches = []

    inv_import = _normalize_code_compare_value(_get_row_value(imported_row, "cod_inventario"))
    inv_db = _normalize_code_compare_value(db_summary.get("cod_inventario"))
    if inv_import and inv_import == inv_db:
        matches.append("cod_inventario")

    esbye_import = _normalize_code_compare_value(_get_row_value(imported_row, "cod_esbye"))
    esbye_db = _normalize_code_compare_value(db_summary.get("cod_esbye"))
    if esbye_import and esbye_import == esbye_db:
        matches.append("cod_esbye")

    serie_import = _normalize_compare_value(_get_row_value(imported_row, "serie"))
    serie_db = _normalize_compare_value(db_summary.get("serie"))
    if serie_import and serie_import == serie_db:
        matches.append("serie")

    modelo_import = _normalize_compare_value(_get_row_value(imported_row, "modelo"))
    modelo_db = _normalize_compare_value(db_summary.get("modelo"))
    if modelo_import and modelo_import == modelo_db:
        matches.append("modelo")

    return matches


def _extract_compact_aula_code(text):
    normalized = _normalize_text(text)
    if not normalized:
        return ""
    m = re.search(r"(\d+[a-z]\s*-?\s*\d+)", normalized)
    if not m:
        return ""
    return re.sub(r"\s+", "", m.group(1))


def _extract_floor_hint(text):
    normalized = _normalize_text(text)
    if "planta baja" in normalized:
        return "planta baja"
    if "primer" in normalized and "piso" in normalized:
        return "primer piso"
    if "segundo" in normalized and "piso" in normalized:
        return "segundo piso"
    if "tercer" in normalized and "piso" in normalized:
        return "tercer piso"
    return ""


def _format_floor_hint_for_message(floor_hint, target_text):
    base = str(floor_hint or "").strip()
    normalized_target = _normalize_text(target_text)
    if not base:
        return ""
    if "planta baja" in base:
        return "planta baja"
    if "segundo piso" in base and "alto" in normalized_target:
        return "segundo piso alto"
    if "segundo piso" in base and "bajo" in normalized_target:
        return "segundo piso bajo"
    if "primer piso" in base and "alto" in normalized_target:
        return "primer piso alto"
    if "primer piso" in base and "bajo" in normalized_target:
        return "primer piso bajo"
    if "tercer piso" in base and "alto" in normalized_target:
        return "tercer piso alto"
    if "tercer piso" in base and "bajo" in normalized_target:
        return "tercer piso bajo"
    return base


def _extract_location_kind(text):
    normalized = _normalize_text(text)
    if not normalized:
        return ""
    if "pasillo" in normalized:
        return "pasillo"
    if "aula" in normalized or _extract_compact_aula_code(normalized):
        return "aula"
    return ""


def _guess_area_match_from_location_text(location_text):
    target = _normalize_text(location_text)
    if not target:
        return None

    target_tokens = [token for token in re.split(r"\s+|/|-", target) if token]
    target_aula_code = _extract_compact_aula_code(target)
    floor_hint = _extract_floor_hint(target)
    target_kind = _extract_location_kind(target)

    db = get_db()
    rows = db.execute(
        """
        SELECT a.id, a.nombre AS area_nombre, p.nombre AS piso_nombre, b.nombre AS bloque_nombre
        FROM areas a
        JOIN pisos p ON p.id = a.piso_id
        JOIN bloques b ON b.id = p.bloque_id
        """
    ).fetchall()

    candidates = []
    pasillo_candidates = []
    explicit_block_letter = ""
    explicit_block_match = re.search(r"\bbloque\s+([a-z])\b", target)
    if explicit_block_match:
        explicit_block_letter = explicit_block_match.group(1)

    for row in rows:
        area_name_raw = str(row["area_nombre"] or "").strip()
        floor_name_raw = str(row["piso_nombre"] or "").strip()
        block_name_raw = str(row["bloque_nombre"] or "").strip()

        area_name = _normalize_text(area_name_raw)
        floor_name = _normalize_text(floor_name_raw)
        block_name = _normalize_text(block_name_raw)
        full_name = f"{block_name} {floor_name} {area_name}".strip()

        score = 0
        if target == area_name or target == full_name:
            score += 30
        if target in full_name:
            score += 10

        area_code = _extract_compact_aula_code(area_name)
        if target_aula_code and area_code and target_aula_code == area_code:
            score += 35

        area_tokens = [token for token in re.split(r"\s+|/|-", full_name) if token]
        token_hits = 0
        for token in target_tokens:
            if len(token) < 3:
                continue
            if any(token in at or at in token for at in area_tokens):
                token_hits += 1
        score += min(token_hits * 2, 16)

        if floor_hint and floor_hint in floor_name:
            score += 8

        if "alto" in target:
            if "alto" in floor_name:
                score += 4
            elif "bajo" in floor_name:
                score -= 3
        if "bajo" in target:
            if "bajo" in floor_name:
                score += 4
            elif "alto" in floor_name:
                score -= 3

        block_letter_match = re.search(r"\d+([a-z])", target_aula_code or "")
        if block_letter_match:
            block_letter = block_letter_match.group(1)
            if f"bloque {block_letter}" in block_name:
                score += 10

        if explicit_block_letter:
            if f"bloque {explicit_block_letter}" in block_name:
                score += 12
            else:
                score -= 16

        # Si el texto fuente pide PASILLO, evitamos caer en aulas del mismo piso por similitud de tokens.
        if target_kind == "pasillo":
            if "pasillo" in area_name:
                score += 22
            else:
                score -= 20
        elif target_kind == "aula":
            if "aula" in area_name or area_code:
                score += 8

        candidate = {
            "id": int(row["id"]),
            "display": " / ".join(part for part in [block_name_raw, floor_name_raw, area_name_raw] if part),
            "area_name": area_name,
            "floor_name": floor_name,
            "score": score,
        }
        candidates.append(candidate)
        if "pasillo" in area_name:
            pasillo_candidates.append(candidate)

    if not candidates:
        return None

    ranked_candidates = sorted(candidates, key=lambda c: c["score"], reverse=True)
    best_candidate = ranked_candidates[0]
    best_score = best_candidate["score"]
    best_id = best_candidate["id"]
    best_display = best_candidate["display"]

    # Si la diferencia con el segundo mejor es muy baja y no hay codigo de aula exacto,
    # evitamos asignar una ubicacion equivocada por similitud ambigua.
    second_best_score = ranked_candidates[1]["score"] if len(ranked_candidates) > 1 else -999
    margin = best_score - second_best_score
    has_exact_aula_code = False
    if target_aula_code:
        best_area_code = _extract_compact_aula_code(best_candidate.get("area_name"))
        has_exact_aula_code = bool(best_area_code and best_area_code == target_aula_code)

    if target_kind == "pasillo":
        # Para entradas tipo PASILLO nunca tomamos aulas como fallback.
        if not pasillo_candidates:
            return None

        best_pasillo = max(
            pasillo_candidates,
            key=lambda c: c["score"] + (5 if floor_hint and floor_hint in c["floor_name"] else 0),
        )
        best_id = best_pasillo["id"]
        best_display = best_pasillo["display"]
        best_score = best_pasillo["score"]

    if best_id is None or best_score < 10:
        return None

    if not has_exact_aula_code and margin <= 1:
        return None

    warning = ""
    if target_kind == "pasillo":
        floor_candidates = [c for c in candidates if floor_hint and floor_hint in c["floor_name"]]
        floor_has_pasillo = any("pasillo" in c["area_name"] for c in floor_candidates)
        floor_exists = bool(floor_candidates)

        if floor_exists and not floor_has_pasillo:
            floor_label = _format_floor_hint_for_message(floor_hint, target)
            warning = f"La ubicación original indica {floor_label}, pero no hay un pasillo registrado en ese piso."

    return {
        "area_id": best_id,
        "display": best_display,
        "warning": warning,
    }


def _guess_area_id_from_location_text(location_text):
    match = _resolve_area_match_with_inventory_method(location_text)
    if not match:
        return None
    return match.get("area_id")


def _resolve_area_match_with_inventory_method(location_text):
    raw = str(location_text or "").strip()
    if not raw:
        return None

    # Resolver unico y mas conservador para evitar aproximaciones incorrectas.
    guessed = _guess_area_match_from_location_text(raw)
    if not guessed or guessed.get("area_id") is None:
        return None

    target_norm = _normalize_text(raw)
    floor_hint = _extract_floor_hint(target_norm)
    target_kind = _extract_location_kind(target_norm)
    display_norm = _normalize_text(guessed.get("display") or "")

    # Si se indica piso explicitamente para PASILLO/AULA, exige que el match respete ese piso.
    if floor_hint and target_kind in {"pasillo", "aula"} and floor_hint not in display_norm:
        return None

    return guessed


def _analyze_chunk_against_inventory(chunk_rows):
    db = get_db()
    db_rows = db.execute(
        """
        SELECT
            id,
            item_numero,
            cod_inventario,
            cod_esbye,
            cuenta,
            cantidad,
            descripcion,
            ubicacion,
            marca,
            modelo,
            serie,
            estado,
            usuario_final,
            fecha_adquisicion,
            valor,
            observacion
        FROM inventario_items
        """
    ).fetchall()

    exact_index = {}
    inv_code_index = {}
    esbye_code_index = {}
    for row in db_rows:
        summary = _inventory_summary_item(row)
        normalized_db_row = _normalize_row_for_exact_compare(dict(row))
        row_key = _build_row_key(normalized_db_row)
        if not _is_empty_row_key(row_key):
            exact_index.setdefault(row_key, []).append(summary)

        inv_code = _normalize_code_compare_value(normalized_db_row.get("cod_inventario"))
        if inv_code:
            inv_code_index.setdefault(inv_code, []).append(summary)

        esbye_code = _normalize_code_compare_value(normalized_db_row.get("cod_esbye"))
        if esbye_code:
            esbye_code_index.setdefault(esbye_code, []).append(summary)

    analyzed = []
    for row in chunk_rows:
        row_data = dict(row.get("data") or {})
        compare_row_data = _normalize_row_for_exact_compare(row_data)
        exact_matches = []
        similar_matches = []

        row_key = _build_row_key(compare_row_data)
        if not _is_empty_row_key(row_key):
            exact_matches = exact_index.get(row_key, [])

        inv_code = _normalize_code_compare_value(compare_row_data.get("cod_inventario"))
        if inv_code:
            similar_matches.extend(inv_code_index.get(inv_code, []))

        esbye_code = _normalize_code_compare_value(compare_row_data.get("cod_esbye"))
        if esbye_code:
            similar_matches.extend(esbye_code_index.get(esbye_code, []))

        # Elimina duplicados por id dentro de similar/exact.
        seen_ids = set()
        dedup_similar = []
        for item in similar_matches:
            item_id = item.get("id")
            if item_id in seen_ids:
                continue
            seen_ids.add(item_id)
            dedup_similar.append(item)
        similar_matches = dedup_similar

        status = "normal"
        if exact_matches:
            status = "exact"
        elif similar_matches:
            status = "similar"

        exact_with_fields = []
        for item in exact_matches:
            enriched = dict(item)
            enriched["match_fields"] = _build_match_fields(compare_row_data, item)
            exact_with_fields.append(enriched)

        similar_with_fields = []
        for item in similar_matches:
            enriched = dict(item)
            enriched["match_fields"] = _build_match_fields(compare_row_data, item)
            similar_with_fields.append(enriched)

        analyzed.append(
            {
                "row_index": row.get("row_index"),
                "data": compare_row_data,
                "status": status,
                "exact_matches": exact_with_fields,
                "similar_matches": similar_with_fields,
            }
        )

    return analyzed


def _build_mapped_chunk(data_rows, mapping, start_index, chunk_size, rows_override=None):
    if isinstance(rows_override, list) and rows_override:
        normalized_rows = []
        for idx, row_data in enumerate(rows_override):
            if not isinstance(row_data, dict):
                continue
            clean_data = {}
            for field, value in row_data.items():
                canonical = str(field or "").strip()
                if canonical in ALLOWED_INVENTORY_FIELDS and canonical != "item_numero":
                    clean_data[canonical] = value
            _normalize_codes_in_row_data(clean_data)
            if not _row_has_meaningful_import_data(clean_data):
                continue
            normalized_rows.append(
                {
                    "row_index": start_index + idx,
                    "data": clean_data,
                }
            )
        return normalized_rows

    chunk = []
    end = min(start_index + chunk_size, len(data_rows))
    for idx in range(start_index, end):
        row = data_rows[idx]
        row_dict = {}
        for col_idx, cell in enumerate(row):
            canonical = str(mapping.get(str(col_idx)) or "").strip()
            if not canonical or canonical not in ALLOWED_INVENTORY_FIELDS or canonical == "item_numero":
                continue
            value = str(cell).strip() if cell is not None else None
            if value in (None, ""):
                continue

            if canonical in ("fecha_adquisicion", "fecha_adquisicion_esbye"):
                date_time_match = re.fullmatch(r"(\d{4}-\d{2}-\d{2})[ T]\d{2}:\d{2}:\d{2}(?:\.\d+)?", value)
                if date_time_match:
                    value = date_time_match.group(1)

            existing_value = row_dict.get(canonical)

            # Si varias columnas se mapean al mismo campo, no pisar un valor util con vacio.
            if existing_value not in (None, "") and value in (None, ""):
                continue
            if existing_value not in (None, "") and value not in (None, ""):
                if canonical in ("observacion", "observacion_esbye") and str(value) != str(existing_value):
                    row_dict[canonical] = f"{existing_value} | {value}"
                continue

            row_dict[canonical] = value
        if row_dict:
            _normalize_codes_in_row_data(row_dict)
            if _row_has_meaningful_import_data(row_dict):
                chunk.append({"row_index": idx, "data": row_dict})
    return chunk


def _clean_old_excel_imports(max_age_seconds=3600):
    now = time.time()
    try:
        for fname in os.listdir(_EXCEL_IMPORT_DIR):
            if not fname.endswith(".xlsx"):
                continue
            fpath = os.path.join(_EXCEL_IMPORT_DIR, fname)
            try:
                if os.path.isfile(fpath) and (now - os.path.getmtime(fpath)) > max_age_seconds:
                    os.remove(fpath)
            except Exception:
                pass
    except Exception:
        pass


def _safe_import_session_path(session_id):
    if not _UUID_RE.fullmatch(str(session_id or "")):
        return None
    safe_dir = os.path.abspath(_EXCEL_IMPORT_DIR)
    expected_name = f"{session_id}.xlsx"
    try:
        for entry in os.scandir(safe_dir):
            if entry.is_file() and entry.name == expected_name:
                return entry.path
    except OSError:
        pass
    return None


@inventory_bp.get("/api/inventario")
def api_list_inventario():
    filters = {
        "bloque_id": request.args.get("bloque_id", type=int),
        "piso_id": request.args.get("piso_id", type=int),
        "area_id": request.args.get("area_id", type=int),
        "search": request.args.get("search", type=str),
    }
    sort_direction = request.args.get("order", default="asc", type=str)
    page = request.args.get("page", default=1, type=int)
    per_page = request.args.get("per_page", default=50, type=int)
    # Quitamos el límite de 500 para permitir extracción completa en actas
    per_page = max(1, per_page)
    include_traspaso_acta_id = request.args.get("include_traspaso_acta_id", type=int)

    result = list_inventory_items_paginated(
        filters=filters,
        sort_direction=sort_direction,
        page=page,
        per_page=per_page,
        include_traspaso_acta_id=include_traspaso_acta_id
    )
    return jsonify(
        {
            "data": result["items"],
            "pagination": {
                "page": result["page"],
                "per_page": result["per_page"],
                "total": result["total"],
                "total_pages": result["total_pages"],
            },
        }
    )


@inventory_bp.get("/api/inventario/bajas")
def api_list_inventario_bajas():
    db = get_db()
    rows = db.execute(
        """
        SELECT i.*
        FROM inventario_items i
        WHERE LOWER(COALESCE(i.procedencia, '')) LIKE 'acta de baja %'
        ORDER BY i.item_numero ASC, i.id ASC
        """
    ).fetchall()
    data = [dict(row) for row in rows]
    return jsonify({"data": data, "total": len(data)})


@inventory_bp.get("/api/inventario/<int:item_id>")
def api_get_inventario(item_id):
    item = get_inventory_item(item_id)
    if not item:
        return jsonify({"error": "Elemento no encontrado."}), 404
    return jsonify({"data": item})


@inventory_bp.post("/api/inventario")
def api_create_inventario():
    payload = request.get_json(silent=True) or {}
    force_duplicate = bool(payload.get("force_duplicate"))
    duplicate_items = find_inventory_code_duplicates(
        cod_inventario=payload.get("cod_inventario"),
        cod_esbye=payload.get("cod_esbye"),
        limit=50,
    )
    if duplicate_items and not force_duplicate:
        return jsonify(
            {
                "error": "Código repetido detectado. Confirma si deseas agregarlo de todas formas.",
                "duplicates": duplicate_items,
            }
        ), 409
    try:
        item_id = create_inventory_item(payload)
    except sqlite3.IntegrityError as error:
        duplicate_items = find_inventory_code_duplicates(
            cod_inventario=payload.get("cod_inventario"),
            cod_esbye=payload.get("cod_esbye"),
            limit=50,
        )
        return jsonify(
            {
                "error": f"No se pudo guardar por un valor duplicado: {error}",
                "duplicates": duplicate_items,
            }
        ), 409
    item = get_inventory_item(item_id)
    return jsonify({"data": item}), 201


@inventory_bp.patch("/api/inventario/<int:item_id>")
def api_update_inventario(item_id):
    payload = request.get_json(silent=True) or {}
    force_duplicate = bool(payload.get("force_duplicate"))
    if "cod_inventario" in payload or "cod_esbye" in payload:
        duplicate_items = find_inventory_code_duplicates(
            cod_inventario=payload.get("cod_inventario") if "cod_inventario" in payload else None,
            cod_esbye=payload.get("cod_esbye") if "cod_esbye" in payload else None,
            limit=50,
            exclude_item_id=item_id,
        )
        if duplicate_items and not force_duplicate:
            return jsonify(
                {
                    "error": "Código repetido detectado. Confirma si deseas guardar el cambio de todas formas.",
                    "duplicates": duplicate_items,
                }
            ), 409
    try:
        ok = update_inventory_item(item_id, payload)
    except sqlite3.IntegrityError as error:
        duplicate_items = find_inventory_code_duplicates(
            cod_inventario=payload.get("cod_inventario") if "cod_inventario" in payload else None,
            cod_esbye=payload.get("cod_esbye") if "cod_esbye" in payload else None,
            limit=50,
            exclude_item_id=item_id,
        )
        return jsonify(
            {
                "error": f"No se pudo actualizar por un valor duplicado: {error}",
                "duplicates": duplicate_items,
            }
        ), 409
    if not ok:
        return jsonify({"error": "Elemento no encontrado."}), 404
    item = get_inventory_item(item_id)
    return jsonify({"data": item})


@inventory_bp.delete("/api/inventario/<int:item_id>")
def api_delete_inventario(item_id):
    ok = delete_inventory_item(item_id)
    if not ok:
        return jsonify({"error": "Elemento no encontrado."}), 404
    return jsonify({"success": True})


@inventory_bp.post("/api/inventario/vaciar-temporal")
def api_clear_inventario_temporal():
    payload = request.get_json(silent=True) or {}
    confirm_text = str(payload.get("confirm_text") or "").strip().upper()
    if confirm_text != "VACIAR TODO":
        return jsonify({"error": "Confirmación inválida. Debes enviar 'VACIAR TODO'."}), 400

    result = clear_inventory_items(reset_sequence=True)
    return jsonify({"success": True, "deleted": result["deleted"]})


@inventory_bp.post("/api/inventario/pegar")
def api_paste_inventario():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "Debe enviar una lista de filas para pegar."}), 400

    area_id_raw = payload.get("area_id")
    try:
        area_id = int(area_id_raw) if area_id_raw not in (None, "", "null") else None
    except (TypeError, ValueError):
        area_id = None

    rows_as_dicts = []
    missing_areas = set()
    location_warnings = []
    for raw_row in rows:
        if not isinstance(raw_row, list):
            continue

        row_data = _map_paste_row_best_effort(raw_row)

        if not row_data:
            continue

        if not row_data.get("area_id"):
            guessed = _resolve_area_match_with_inventory_method(row_data.get("ubicacion"))
            if guessed and guessed.get("area_id") is not None:
                row_data["area_id"] = guessed.get("area_id")
                if guessed.get("warning"):
                    location_warnings.append(
                        {
                            "row_index": len(rows_as_dicts),
                            "ubicacion": str(row_data.get("ubicacion") or "").strip(),
                            "warning": guessed.get("warning"),
                            "suggested_area": guessed.get("display"),
                        }
                    )
        if area_id is not None and not row_data.get("area_id"):
            row_data["area_id"] = area_id

        if row_data.get("ubicacion") and not row_data.get("area_id"):
            missing_areas.add(str(row_data.get("ubicacion") or "").strip())

        # Reutiliza exactamente la misma normalizacion de usuario_final del importador Excel.
        raw_usuario = str(row_data.get("usuario_final") or "").strip()
        if raw_usuario:
            row_data["usuario_final"] = resolve_or_create_personal_name(
                raw_usuario,
                create_if_missing=True,
            )

        rows_as_dicts.append(row_data)

    missing_areas.discard("")
    if missing_areas:
        return jsonify(
            {
                "error": "Se encontraron ubicaciones/areas que no existen en la configuración.",
                "code": "area_not_found",
                "missing_areas": sorted(missing_areas),
            }
        ), 409

    try:
        result = bulk_insert_inventory_dicts(
            rows_as_dicts,
            area_id=area_id,
            procedencia_default=_build_bulk_excel_procedencia_text(),
        )
    except sqlite3.IntegrityError as error:
        return jsonify({"error": f"Error al pegar datos por duplicado: {error}"}), 409
    return jsonify(
        {
            "inserted": int(result.get("inserted") or 0),
            "skipped": int(result.get("skipped") or 0),
            "location_warnings": location_warnings,
        }
    ), 201


@inventory_bp.get("/api/inventario/search-diagnostics")
def api_inventory_search_diagnostics():
    search_text = request.args.get("search", default="", type=str)
    data = get_inventory_search_diagnostics(search_text)
    return jsonify({"data": data})


@inventory_bp.get("/api/inventario/duplicados")
def api_inventory_duplicates():
    cod_inventario = request.args.get("cod_inventario", default="", type=str)
    cod_esbye = request.args.get("cod_esbye", default="", type=str)
    duplicates = find_inventory_code_duplicates(
        cod_inventario=cod_inventario,
        cod_esbye=cod_esbye,
        limit=50,
    )
    return jsonify({"success": True, "duplicates": duplicates})


@inventory_bp.get("/api/inventario/resolver-ubicacion")
def api_inventory_resolve_location():
    raw_text = request.args.get("texto", default="", type=str)
    text = str(raw_text or "").strip()
    if not text:
        return jsonify({"success": False, "error": "Parámetro 'texto' es obligatorio."}), 400

    guessed = _resolve_area_match_with_inventory_method(text)
    if not guessed or guessed.get("area_id") is None:
        return jsonify({"success": True, "match": None})

    return jsonify(
        {
            "success": True,
            "match": {
                "area_id": int(guessed.get("area_id")),
                "display": str(guessed.get("display") or "").strip(),
                "warning": str(guessed.get("warning") or "").strip(),
            },
        }
    )


@inventory_bp.get("/api/preferencias")
def api_get_preferencias():
    return jsonify({"data": get_user_preferences(DEFAULT_USER_KEY)})


@inventory_bp.patch("/api/preferencias")
def api_set_preferencias():
    payload = request.get_json(silent=True) or {}
    pref_key = (payload.get("pref_key") or "").strip()
    if not pref_key:
        return jsonify({"error": "pref_key es obligatorio."}), 400
    set_user_preference(DEFAULT_USER_KEY, pref_key, payload.get("pref_value"))
    return jsonify({"success": True})


@inventory_bp.get("/api/column-mappings")
def api_get_column_mappings():
    return jsonify({"data": get_column_mappings()})


@inventory_bp.patch("/api/column-mappings")
def api_put_column_mappings():
    payload = request.get_json(silent=True) or {}
    mappings = payload.get("mappings") or []
    if not isinstance(mappings, list):
        return jsonify({"error": "mappings debe ser una lista."}), 400
    replace_column_mappings(mappings)
    return jsonify({"success": True})


@inventory_bp.post("/api/inventario/previsualizar-excel")
def api_previsualizar_excel():
    _clean_old_excel_imports()

    if "file" not in request.files:
        return jsonify({"error": "No se recibió ningún archivo."}), 400

    file = request.files["file"]
    filename = (file.filename or "").strip().lower()
    if not filename.endswith(".xlsx"):
        return jsonify({"error": "Solo se aceptan archivos .xlsx"}), 400

    content = file.read()
    if len(content) > _MAX_EXCEL_FILE_BYTES:
        return jsonify({"error": "El archivo supera el límite de 10 MB."}), 400
    if not content:
        return jsonify({"error": "El archivo está vacío."}), 400

    session_id = str(uuid.uuid4())
    temp_path = os.path.join(_EXCEL_IMPORT_DIR, f"{session_id}.xlsx")
    with open(temp_path, "wb") as f:
        f.write(content)

    try:
        parsed = _load_excel_with_detected_headers(temp_path)
        headers = parsed["headers"]
        all_data_rows = parsed["data_rows"]
        preview_rows = all_data_rows[:_MAX_EXCEL_PREVIEW_ROWS]
        total_rows = len(all_data_rows)
        suggested_mapping = [_header_to_canonical_field(header) for header in headers]
        suggested_mapping = _normalize_suggested_mapping_by_repetition(headers, suggested_mapping)
    except Exception:
        logger.exception("Error reading uploaded Excel file")
        try:
            os.remove(temp_path)
        except Exception:
            pass
        return jsonify({"error": "No se pudo leer el archivo Excel. Verifica que sea un .xlsx válido y no esté dañado."}), 400

    if not headers:
        try:
            os.remove(temp_path)
        except Exception:
            pass
        return jsonify({"error": "El archivo no tiene encabezados en la primera fila."}), 400

    return jsonify(
        {
            "session_id": session_id,
            "headers": headers,
            "preview_rows": preview_rows,
            "total_rows": total_rows,
            "header_row_index": int(parsed["header_row_index"]),
            "suggested_mapping": suggested_mapping,
        }
    )


@inventory_bp.post("/api/inventario/confirmar-importacion")
def api_confirmar_importacion():
    payload = request.get_json(silent=True) or {}
    session_id = str(payload.get("session_id") or "").strip()
    mapping = payload.get("mapping") or {}
    area_id_raw = payload.get("area_id")
    start_index = max(_coerce_int(payload.get("start_index"), 0), 0)
    chunk_size = _coerce_int(payload.get("chunk_size"), _DEFAULT_IMPORT_CHUNK_SIZE)
    if chunk_size <= 0:
        chunk_size = _DEFAULT_IMPORT_CHUNK_SIZE
    chunk_size = min(chunk_size, 100)
    validate_only = bool(payload.get("validate_only"))
    force_duplicate = bool(payload.get("force_duplicate"))
    rows_override = payload.get("rows")

    if not session_id:
        return jsonify({"error": "Falta session_id."}), 400

    temp_path = _safe_import_session_path(session_id)
    if temp_path is None:
        return jsonify({"error": "La sesión de importación ha expirado. Por favor sube el archivo nuevamente."}), 410

    try:
        area_id = int(area_id_raw) if area_id_raw not in (None, "", "null") else None
    except (TypeError, ValueError):
        area_id = None

    try:
        parsed = _load_excel_with_detected_headers(temp_path)
        data_rows = parsed["data_rows"]
        total_rows = len(data_rows)
        mapped_chunk = _build_mapped_chunk(data_rows, mapping, start_index, chunk_size, rows_override=rows_override)
        analyzed_chunk = _analyze_chunk_against_inventory(mapped_chunk)
    except Exception:
        logger.exception("Error reading Excel file during import")
        return jsonify({"error": "Error al leer el archivo para importar. Por favor sube el archivo nuevamente."}), 500

    if not mapped_chunk:
        return jsonify({"error": "No se encontraron filas con datos para este bloque."}), 400

    exact_rows = [row for row in analyzed_chunk if row["status"] == "exact"]
    similar_rows = [row for row in analyzed_chunk if row["status"] == "similar"]
    has_conflicts = bool(exact_rows or similar_rows)

    if validate_only:
        return jsonify(
            {
                "success": True,
                "total_rows": total_rows,
                "start_index": start_index,
                "chunk_size": chunk_size,
                "has_more": (start_index + chunk_size) < total_rows,
                "next_start_index": start_index + chunk_size,
                "rows": analyzed_chunk,
                "summary": {
                    "exact": len(exact_rows),
                    "similar": len(similar_rows),
                    "normal": len(analyzed_chunk) - len(exact_rows) - len(similar_rows),
                },
            }
        )

    if has_conflicts and not force_duplicate:
        return jsonify(
            {
                "error": "Se detectaron filas ya registradas o similares. Confirma si deseas guardar de todas formas.",
                "rows": analyzed_chunk,
                "summary": {
                    "exact": len(exact_rows),
                    "similar": len(similar_rows),
                    "normal": len(analyzed_chunk) - len(exact_rows) - len(similar_rows),
                },
            }
        ), 409

    rows_as_dicts = []
    invalid_locations = []
    location_warnings = []
    for row in analyzed_chunk:
        row_data = dict(row.get("data") or {})
        if not _row_has_meaningful_import_data(row_data):
            continue

        raw_area_id = row_data.get("area_id")
        has_explicit_area = False
        if raw_area_id not in (None, "", "null"):
            try:
                row_data["area_id"] = int(raw_area_id)
                has_explicit_area = bool(row_data["area_id"])
            except (TypeError, ValueError):
                row_data["area_id"] = None

        raw_location = str(row_data.get("ubicacion") or "").strip()
        if raw_location and not has_explicit_area:
            guessed = _resolve_area_match_with_inventory_method(raw_location)
            if guessed and guessed.get("area_id") is not None:
                row_data["area_id"] = guessed.get("area_id")
                if guessed.get("warning"):
                    location_warnings.append(
                        {
                            "row_index": row.get("row_index"),
                            "ubicacion": raw_location,
                            "warning": guessed.get("warning"),
                            "suggested_area": guessed.get("display"),
                        }
                    )
            else:
                invalid_locations.append(
                    {
                        "row_index": row.get("row_index"),
                        "ubicacion": raw_location,
                        "descripcion": row_data.get("descripcion"),
                        "cod_inventario": row_data.get("cod_inventario"),
                        "cod_esbye": row_data.get("cod_esbye"),
                    }
                )

        if area_id is not None and not row_data.get("area_id"):
            row_data["area_id"] = area_id

        raw_usuario = str(row_data.get("usuario_final") or "").strip()
        if raw_usuario:
            row_data["usuario_final"] = resolve_or_create_personal_name(
                raw_usuario,
                create_if_missing=True,
            )

        rows_as_dicts.append(row_data)

    if invalid_locations:
        return jsonify(
            {
                "error": "Se encontraron ubicaciones que no existen en la configuración.",
                "error_code": "invalid_locations",
                "invalid_locations": invalid_locations,
            }
        ), 422

    try:
        result = bulk_insert_inventory_dicts(rows_as_dicts, area_id=area_id)
    except sqlite3.IntegrityError:
        logger.exception("Integrity error during Excel bulk import")
        return jsonify({"error": "Error de integridad al importar. Puede haber datos con formato inválido."}), 409

    # Solo borramos el archivo temporal cuando ya no quedan filas por procesar.
    has_more = (start_index + chunk_size) < total_rows
    if not has_more:
        try:
            os.remove(temp_path)
        except Exception:
            pass

    return jsonify(
        {
            "success": True,
            "inserted": result["inserted"],
            "skipped": result["skipped"],
            "location_warnings": location_warnings,
            "has_more": has_more,
            "next_start_index": start_index + chunk_size,
            "total_rows": total_rows,
            "summary": {
                "exact": len(exact_rows),
                "similar": len(similar_rows),
                "normal": len(analyzed_chunk) - len(exact_rows) - len(similar_rows),
            },
        }
    ), 201


@inventory_bp.post("/api/inventario/excel-a-filas-recepcion")
def api_excel_to_recepcion_rows():
    payload = request.get_json(silent=True) or {}
    session_id = str(payload.get("session_id") or "").strip()
    forced_location = str(payload.get("forced_location") or "").strip()
    forced_area_id_raw = payload.get("forced_area_id")
    mapping = payload.get("mapping") or {}
    validate_only = bool(payload.get("validate_only"))
    force_duplicate = bool(payload.get("force_duplicate"))
    start_index = max(_coerce_int(payload.get("start_index"), 0), 0)
    chunk_size = _coerce_int(payload.get("chunk_size"), 20)
    if chunk_size <= 0:
        chunk_size = 20
    chunk_size = min(chunk_size, 200)

    if not session_id:
        return jsonify({"error": "Falta session_id."}), 400
    if not forced_location:
        return jsonify({"error": "Debes seleccionar Bloque, Piso y Área antes de importar a recepción."}), 400

    temp_path = _safe_import_session_path(session_id)
    if temp_path is None:
        return jsonify({"error": "La sesión de importación ha expirado. Sube el archivo nuevamente."}), 410

    try:
        forced_area_id = int(forced_area_id_raw) if forced_area_id_raw not in (None, "", "null") else None
    except (TypeError, ValueError):
        forced_area_id = None

    try:
        parsed = _load_excel_with_detected_headers(temp_path)
        data_rows = parsed.get("data_rows") or []

        # Si no se envía mapeo manual, usa el sugerido por encabezados.
        if not isinstance(mapping, dict) or not mapping:
            headers = parsed.get("headers") or []
            mapping = {}
            for idx, header in enumerate(headers):
                mapping[str(idx)] = _header_to_canonical_field(header)

        mapped_rows = _build_mapped_chunk(
            data_rows,
            mapping,
            start_index=start_index,
            chunk_size=chunk_size,
        )
        analyzed_rows = _analyze_chunk_against_inventory(mapped_rows)
    except Exception:
        logger.exception("Error parsing Excel for recepcion temp import")
        return jsonify({"error": "No se pudo leer el Excel para recepción."}), 500

    exact_rows = [row for row in analyzed_rows if row.get("status") == "exact"]
    similar_rows = [row for row in analyzed_rows if row.get("status") == "similar"]
    has_conflicts = bool(exact_rows or similar_rows)

    rows = []
    skipped = 0
    for row in analyzed_rows:
        row_data = dict(row.get("data") or {})
        if not row_data:
            skipped += 1
            continue

        # Ubicación forzada por el área seleccionada en acta de recepción.
        row_data["ubicacion"] = forced_location
        if forced_area_id is not None:
            row_data["area_id"] = forced_area_id

        raw_usuario = str(row_data.get("usuario_final") or "").strip()
        if raw_usuario:
            row_data["usuario_final"] = resolve_or_create_personal_name(
                raw_usuario,
                create_if_missing=False,
            )

        rows.append(row_data)

    total_data_rows = len(data_rows)
    has_more = (start_index + chunk_size) < total_data_rows
    summary = {
        "exact": len(exact_rows),
        "similar": len(similar_rows),
        "normal": len(analyzed_rows) - len(exact_rows) - len(similar_rows),
    }

    if validate_only:
        return jsonify(
            {
                "success": True,
                "rows": rows,
                "total_rows": total_data_rows,
                "chunk_rows": len(rows),
                "skipped": skipped,
                "start_index": start_index,
                "chunk_size": chunk_size,
                "next_start_index": start_index + chunk_size,
                "has_more": has_more,
                "summary": summary,
                "analyzed_rows": analyzed_rows,
            }
        )

    if has_conflicts and not force_duplicate:
        return jsonify(
            {
                "error": "Se detectaron filas ya registradas o similares en este bloque.",
                "rows": rows,
                "total_rows": total_data_rows,
                "chunk_rows": len(rows),
                "skipped": skipped,
                "start_index": start_index,
                "chunk_size": chunk_size,
                "next_start_index": start_index + chunk_size,
                "has_more": has_more,
                "summary": summary,
                "analyzed_rows": analyzed_rows,
            }
        ), 409

    if not has_more:
        try:
            os.remove(temp_path)
        except Exception:
            pass

    return jsonify(
        {
            "success": True,
            "rows": rows,
            "total_rows": total_data_rows,
            "chunk_rows": len(rows),
            "skipped": skipped,
            "start_index": start_index,
            "chunk_size": chunk_size,
            "next_start_index": start_index + chunk_size,
            "has_more": has_more,
            "summary": summary,
            "analyzed_rows": analyzed_rows,
        }
    )
